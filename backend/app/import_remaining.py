"""
Import remaining products from the Excel file into the 3DJAT database.
Reads 'Jobs' and 'final' sheets, matches by name, inserts missing products.
"""
import os
import sys
import openpyxl
import sqlite3
import re

# Paths
EXCEL_PATH = r'C:\Users\barba\.hermes\desktop-attachments\Prodouct(AutoRecovered).xlsx'
DB_PATH = r'C:\Users\barba\3djat-pricing\data\3djat.db'

def parse_print_time_hours(td):
    """Convert timedelta to hours (float)."""
    if td is None:
        return 0.0
    if hasattr(td, 'total_seconds'):
        return td.total_seconds() / 3600.0
    return 0.0

def strip_product_id_prefix(name):
    """
    Strip product ID prefix like 'KE001. ' from product name.
    Returns (stripped_name, had_prefix).
    """
    m = re.match(r'^[A-Z]{2}\d{3}\.\s*', name)
    if m:
        return name[m.end():].strip(), True
    return name.strip(), False

def determine_category(product_id):
    """Determine category from product ID prefix."""
    if not product_id:
        return ''
    pid = str(product_id).strip().upper()
    for prefix in ['KE', 'CO', 'SK', 'BO', 'PT']:
        if pid.startswith(prefix):
            return prefix
    return ''

def build_final_mapping(wb):
    """Build a dict from the 'final' sheet: product_name -> (persian_name, final_price)."""
    ws = wb['final']
    mapping = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = row[1]
        if name is None or str(name).strip() == '' or str(name).strip() == '0':
            continue
        name = str(name).strip()
        persian = str(row[2]).strip() if row[2] and str(row[2]).strip() != 'None' else None
        final_price = row[5] if row[5] is not None else None
        # Convert final_price to number if possible
        if final_price is not None:
            try:
                final_price = float(final_price)
                if final_price == 0:
                    final_price = None
            except (ValueError, TypeError):
                final_price = None
        if persian or final_price:
            mapping[name] = (persian, final_price)
    return mapping

def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # Build final sheet mapping
    final_map = build_final_mapping(wb)
    print(f"Loaded {len(final_map)} entries from 'final' sheet")
    
    # Connect to DB
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # Get existing product names (for matching)
    cur.execute("SELECT id, name, product_id FROM products")
    existing = cur.fetchall()
    existing_names = {}  # normalized_name -> row
    for row in existing:
        norm = row['name'].strip().lower()
        existing_names[norm] = row
    
    print(f"Existing products in DB: {len(existing)}")
    
    # Build machine name -> id mapping
    cur.execute("SELECT id, name FROM machines")
    machine_map = {}
    for row in cur.fetchall():
        machine_map[row['name'].strip()] = row['id']
    
    # Build material (name, color) -> id mapping (case-insensitive)
    cur.execute("SELECT id, name, color FROM materials")
    material_map = {}  # (name.lower(), color.lower()) -> id
    for row in cur.fetchall():
        key = (row['name'].strip().lower(), row['color'].strip().lower())
        material_map[key] = row['id']
    
    # Read Jobs sheet
    ws = wb['Jobs']
    products_to_insert = []
    skipped = 0
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] is None or str(row[1]).strip() == '':
            continue
        
        excel_name = str(row[1]).strip()
        pid_raw = row[0] if row[0] else ''
        product_id = str(pid_raw).strip() if pid_raw else ''
        
        # Skip #REF! product IDs
        if product_id in ('#REF!', '#VALUE!', '0', ''):
            product_id = ''
        
        qty = row[2] if row[2] else 1
        machine_name = str(row[3]).strip() if row[3] else ''
        material_name = str(row[4]).strip() if row[4] else ''
        color = str(row[5]).strip() if row[5] else ''
        weight = float(row[6]) if row[6] else 0
        support = float(row[7]) if row[7] else 0
        flushed = float(row[8]) if row[8] else 0
        print_time = parse_print_time_hours(row[9])
        post_pro = parse_print_time_hours(row[10])
        extras = float(row[11]) if row[11] else 0
        
        # Determine the display name (strip prefix if present)
        display_name, had_prefix = strip_product_id_prefix(excel_name)
        
        # If we have a product_id and the Excel name has the prefix,
        # use the stripped name as display name
        if product_id and had_prefix:
            name = display_name
        else:
            name = excel_name
        
        # Check if this product already exists in DB
        norm_name = name.strip().lower()
        if norm_name in existing_names:
            skipped += 1
            continue
        
        # Also check the full Excel name (with prefix)
        norm_excel = excel_name.strip().lower()
        if norm_excel in existing_names:
            skipped += 1
            continue
        
        # Determine category
        category = determine_category(product_id)
        
        # Map machine
        machine_id = machine_map.get(machine_name)
        if machine_id is None:
            print(f"  WARNING: Unknown machine '{machine_name}' for product '{name}'")
            continue
        
        # Map material (case-insensitive)
        mat_key = (material_name.lower(), color.lower())
        material_id = material_map.get(mat_key)
        
        # Special case: PLA + Olive Green -> PLA (Olive Green)
        if material_id is None and material_name.lower() == 'pla' and color.lower() == 'olive green':
            material_id = material_map.get(('pla +', 'olive green'))
        
        if material_id is None:
            print(f"  WARNING: Unknown material '{material_name}' color '{color}' for product '{name}'")
            continue
        
        # Get Persian name and final price from final sheet
        persian_name = None
        final_price = None
        # Try matching by Excel name first, then by display name
        if excel_name in final_map:
            persian_name, final_price = final_map[excel_name]
        elif name in final_map:
            persian_name, final_price = final_map[name]
        
        # Use Persian name as the product name if available and the name is Latin
        if persian_name and not any('\u0600' <= c <= '\u06FF' for c in name):
            # Name is Latin-only, use Persian name
            name = persian_name
        
        products_to_insert.append({
            'product_id': product_id,
            'name': name,
            'qty': qty,
            'machine_id': machine_id,
            'material_id': material_id,
            'weight_g': weight,
            'support_g': support,
            'flushed_g': flushed,
            'print_time_hours': print_time,
            'post_pro_hours': post_pro,
            'extras_cost': extras,
            'final_price': final_price,
            'category': category,
            'notes': '',
        })
    
    print(f"\nProducts to insert: {len(products_to_insert)}")
    print(f"Products skipped (already in DB): {skipped}")
    
    # Insert products
    inserted = 0
    for p in products_to_insert:
        cur.execute("""
            INSERT INTO products 
            (product_id, name, qty, machine_id, material_id, weight_g, support_g, 
             flushed_g, print_time_hours, post_pro_hours, extras_cost, final_price, 
             category, notes, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        """, (
            p['product_id'], p['name'], p['qty'], p['machine_id'], p['material_id'],
            p['weight_g'], p['support_g'], p['flushed_g'], p['print_time_hours'],
            p['post_pro_hours'], p['extras_cost'], p['final_price'],
            p['category'], p['notes']
        ))
        inserted += 1
    
    conn.commit()
    
    # Verify
    cur.execute("SELECT COUNT(*) FROM products")
    total = cur.fetchone()[0]
    print(f"\nInserted {inserted} new products")
    print(f"Total products in DB now: {total}")
    
    # Show breakdown by category
    cur.execute("SELECT category, COUNT(*) FROM products GROUP BY category ORDER BY category")
    print("\nBreakdown by category:")
    for row in cur.fetchall():
        cat = row[0] or '(uncategorized)'
        print(f"  {cat}: {row[1]}")
    
    conn.close()
    print("\nDone!")

if __name__ == '__main__':
    main()
