from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import logging
import uuid, os, shutil, io, csv
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

from app.database import get_db
from app.models import Product, Machine, Material, ProductImage, Collection, ProductCollection
from app.repositories.products import ProductRepository, batch_load_machines_and_materials as _batch_load_related
from app.schemas import (
    ProductCreate, ProductUpdate, ProductResponse, ImageReorderRequest,
    CalculateRequest, CalculateResponse,
)
from app.calculator import calculate_product_costs, calculate_product_costs_from_dicts
from app.cache import get_settings_dict
from app.routers.stats import invalidate_stats
from app.routers.auth import require_admin, require_any_role
from app.audit import log_user
from pydantic import BaseModel
from app.services.image import validate_image_bytes as _validate_image_bytes, process_and_save_image as _process_and_save_image

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10MB

router = APIRouter(prefix="/api/v1", tags=["products"])


def _enrich_product(product: Product, db: Session, machines_dict: dict = None, materials_dict: dict = None) -> dict:
    """Build ProductResponse dict with computed costs."""
    # Use pre-fetched dicts when available to avoid per-product DB queries
    if machines_dict is not None and materials_dict is not None:
        mat = materials_dict.get(product.material_id) if product.material_id else None
        mach = machines_dict.get(product.machine_id) if product.machine_id else None
        settings = get_settings_dict(db)
        costs = calculate_product_costs_from_dicts(product, mat, mach, settings)
        machine_name = mach.name if mach else None
        material_name = mat.name if mat else None
        material_color = mat.color if mat else None
    else:
        costs = calculate_product_costs(
            db,
            weight_g=product.weight_g,
            support_g=product.support_g,
            flushed_g=product.flushed_g,
            print_time_hours=product.print_time_hours,
            post_pro_hours=product.post_pro_hours,
            extras_cost=product.extras_cost,
            machine_id=product.machine_id,
            material_id=product.material_id,
        )
        machine_name = product.machine.name if product.machine else None
        material_name = product.material.name if product.material else None
        material_color = product.material.color if product.material else None

    data = {
        "id": product.id,
        "product_id": product.product_id,
        "name": product.name,
        "qty": product.qty,
        "machine_id": product.machine_id,
        "machine_name": machine_name,
        "material_id": product.material_id,
        "material_name": material_name,
        "material_color": material_color,
        "weight_g": product.weight_g,
        "support_g": product.support_g,
        "flushed_g": product.flushed_g,
        "dimension_x": getattr(product, 'dimension_x', None),
        "dimension_y": getattr(product, 'dimension_y', None),
        "dimension_z": getattr(product, 'dimension_z', None),
        "print_time_hours": product.print_time_hours,
        "post_pro_hours": product.post_pro_hours,
        "extras_cost": product.extras_cost,
        "image_url": product.image_url,
        "slug": getattr(product, 'slug', None),
        "tags": getattr(product, 'tags', None),
        "images": [
            {"id": img.id, "image_url": img.image_url, "sort_order": img.sort_order, "is_primary": img.is_primary}
            for img in (product.images if hasattr(product, 'images') and product.images else [])
        ],
        "final_price": product.final_price,
        "category": product.category,
        "notes": product.notes,
        "package_info": product.package_info or "",
        "is_active": product.is_active,
    }
    enriched = {**data, **costs}
    enriched["categories"] = [{"id": c.id, "name": c.name} for c in (product.categories or []) if hasattr(product, "categories")]
    enriched["collections"] = [{"id": c.id, "name": c.name, "slug": c.slug} for c in (product.collections or []) if hasattr(product, "collections")]
    return enriched





# ── Specific routes BEFORE parameterized ──────────────────────────────

@router.get("/products/all", response_model=list[ProductResponse])
def get_all_products(db: Session = Depends(get_db)):
    """Return ALL products including inactive, with computed costs."""
    repo = ProductRepository(db)
    products = repo.get_all(active_only=False)
    machines_dict, materials_dict = _batch_load_related(db)
    return [_enrich_product(p, db, machines_dict, materials_dict) for p in products]


@router.get("/products", response_model=list[ProductResponse])
def get_active_products(db: Session = Depends(get_db)):
    """Return only active products with computed costs."""
    repo = ProductRepository(db)
    products = repo.get_all(active_only=True)
    machines_dict, materials_dict = _batch_load_related(db)
    return [_enrich_product(p, db, machines_dict, materials_dict) for p in products]


# ── Import / Export ─────────────────────────────────────────────────

EXPORT_COLUMNS = [
    "product_id", "name", "category", "material_name", "material_color",
    "weight_g", "support_g", "flushed_g", "print_time_hours",
    "post_pro_hours", "extras_cost", "final_price", "notes", "is_active",
]


@router.get("/products/export")
def export_products(user=Depends(require_admin), db: Session = Depends(get_db)):
    """Export all products as .xlsx file."""
    products = db.query(Product).filter(Product.is_active == True).all()
    machines_dict, materials_dict = _batch_load_related(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header row
    ws.append(EXPORT_COLUMNS)

    for p in products:
        mat = materials_dict.get(p.material_id) if p.material_id else None
        row = [
            p.product_id or "",
            p.name or "",
            p.category or "",
            mat.name if mat else "",
            mat.color if mat else "",
            p.weight_g or 0,
            p.support_g or 0,
            p.flushed_g or 0,
            p.print_time_hours or 0,
            p.post_pro_hours or 0,
            p.extras_cost or 0,
            p.final_price if p.final_price is not None else "",
            p.notes or "",
            p.is_active,
        ]
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_export.xlsx"},
    )


@router.post("/products/import")
async def import_products(
    file: UploadFile = File(...),
    user=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Import products from .xlsx or .csv file. Returns summary of created/updated/errors."""
    content = await file.read()
    filename = file.filename or ""

    # Pre-fetch lookup tables
    machines_dict = {m.name.lower(): m for m in db.query(Machine).filter(Machine.is_active == True).all()}
    materials_dict = {m.name.lower(): m for m in db.query(Material).filter(Material.is_active == True).all()}
    products_by_name = {p.name.lower(): p for p in db.query(Product).all()}

    rows = []
    errors = []
    created = 0
    updated = 0

    try:
        if filename.lower().endswith(".csv"):
            text = content.decode("utf-8-sig")  # Handle BOM
            reader = csv.DictReader(io.StringIO(text))
            for i, row in enumerate(reader, start=2):
                rows.append((i, row))
        elif filename.lower().endswith((".xlsx", ".xls")):
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if not headers:
                raise HTTPException(status_code=400, detail="فایل خالی است")
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
                rows.append((i, row_dict))
            wb.close()
        else:
            raise HTTPException(status_code=400, detail="فرمت فایل پشتیبانی نمی‌شود. فقط .xlsx و .csv")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"خطا در خواندن فایل: {str(e)}")

    for row_num, row in rows:
        name = str(row.get("name", "") or "").strip()
        if not name:
            errors.append({"row": row_num, "error": "نام محصول الزامی است"})
            continue

        # Lookup material by name
        material_name = str(row.get("material_name", "") or "").strip()
        material = materials_dict.get(material_name.lower()) if material_name else None
        if material_name and not material:
            errors.append({"row": row_num, "error": f"ماتریال '{material_name}' یافت نشد"})
            continue

        # Lookup machine by name (from column if present)
        machine_name = str(row.get("machine_name", "") or "").strip()
        machine = machines_dict.get(machine_name.lower()) if machine_name else None
        if machine_name and not machine:
            errors.append({"row": row_num, "error": f"پرینتر '{machine_name}' یافت نشد"})
            continue

        # Parse numeric fields
        def _float(val, default=0.0):
            if val is None or val == "":
                return default
            try:
                return float(val)
            except (ValueError, TypeError):
                return default

        def _bool(val, default=True):
            if val is None or val == "":
                return default
            if isinstance(val, bool):
                return val
            s = str(val).strip().lower()
            return s not in ("false", "0", "no", "غیرفعال")

        product_id_str = str(row.get("product_id", "") or "").strip()
        category = str(row.get("category", "") or "").strip()
        weight_g = _float(row.get("weight_g"))
        support_g = _float(row.get("support_g"))
        flushed_g = _float(row.get("flushed_g"))
        print_time_hours = _float(row.get("print_time_hours"))
        post_pro_hours = _float(row.get("post_pro_hours"))
        extras_cost = _float(row.get("extras_cost"))
        final_price_raw = row.get("final_price")
        final_price = _float(final_price_raw, None) if final_price_raw not in (None, "") else None
        notes = str(row.get("notes", "") or "").strip()
        is_active = _bool(row.get("is_active"))

        # Build data dict
        data = {
            "product_id": product_id_str,
            "name": name,
            "category": category,
            "weight_g": weight_g,
            "support_g": support_g,
            "flushed_g": flushed_g,
            "print_time_hours": print_time_hours,
            "post_pro_hours": post_pro_hours,
            "extras_cost": extras_cost,
            "final_price": final_price,
            "notes": notes,
            "is_active": is_active,
        }
        if material:
            data["material_id"] = material.id
        if machine:
            data["machine_id"] = machine.id

        existing = products_by_name.get(name.lower())
        try:
            if existing:
                repo = ProductRepository(db)
                repo.update(existing.id, data)
                updated += 1
            else:
                repo = ProductRepository(db)
                new_prod = repo.create(data)
                products_by_name[name.lower()] = new_prod
                created += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": row_num, "error": f"خطا در ذخیره: {str(e)}"})
            continue

    return {"created": created, "updated": updated, "errors": errors}


@router.get("/products/{product_id}", response_model=ProductResponse)
def get_product(product_id: int, db: Session = Depends(get_db)):
    """Return single product with computed costs."""
    repo = ProductRepository(db)
    product = repo.get_by_id(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return _enrich_product(product, db)


@router.post("/products", response_model=ProductResponse, status_code=201)
def create_product(product: ProductCreate, user=Depends(require_any_role), db: Session = Depends(get_db)):
    try:
        data = product.model_dump()
        # Pop category_ids & collection_ids before passing to SQLAlchemy model
        category_ids = data.pop("category_ids", None)
        collection_ids = data.pop("collection_ids", None)
        # Auto-generate slug from name if not provided
        if not data.get("slug") and data.get("name"):
            data["slug"] = _slugify(data["name"], data.get("product_id", ""))
        # Ensure slug uniqueness
        if data.get("slug"):
            base = data["slug"]
            i = 1
            while db.query(Product).filter(Product.slug == data["slug"]).first():
                data["slug"] = f"{base}-{i}"
                i += 1
        # Create product without committing yet
        new_prod = Product(**data)
        db.add(new_prod)
        db.flush()  # Get the ID without committing
        
        # Add category associations
        if category_ids:
            from app.models import Category, ProductCategory as PC
            for cat_id in category_ids:
                cat = db.query(Category).filter(Category.id == cat_id).first()
                if cat:
                    db.add(PC(product_id=new_prod.id, category_id=cat_id))

        # Add collection associations
        if collection_ids:
            from app.models import Collection, ProductCollection as PColl
            for coll_id in collection_ids:
                coll = db.query(Collection).filter(Collection.id == coll_id).first()
                if coll:
                    db.add(PColl(product_id=new_prod.id, collection_id=coll_id))
        
        # Commit everything together
        db.commit()
        db.refresh(new_prod)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail="خطا در ایجاد محصول")
    invalidate_stats()
    log_user(user, db, "create", "product", new_prod.id, f"ایجاد محصول «{new_prod.name}»")
    return _enrich_product(new_prod, db)


def _slugify(name: str, product_id: str = "") -> str:
    """Convert a product name to a URL-safe slug.
    Prefers product_id (e.g. KE015) over Persian name which becomes empty when stripped to ASCII.
    """
    import re
    import unicodedata
    # Prefer product_id if it looks like a valid code
    if product_id and re.match(r"^[A-Za-z0-9_-]+$", product_id):
        return product_id.lower()
    # Fallback to name slugification
    nfd = unicodedata.normalize("NFKD", name)
    ascii_name = nfd.encode("ascii", "ignore").decode("ascii").lower()
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_name).strip("-")
    return slug or "product"


@router.put("/products/{product_id}", response_model=ProductResponse)
def update_product(product_id: int, product: ProductUpdate, user=Depends(require_any_role), db: Session = Depends(get_db)):
    repo = ProductRepository(db)
    update_data = product.model_dump(exclude_unset=True)
    category_ids = update_data.pop("category_ids", None)
    collection_ids = update_data.pop("collection_ids", None)
    updated = repo.update(product_id, update_data)
    if not updated:
        raise HTTPException(status_code=404, detail="Product not found")
    if category_ids is not None:
        from app.models import Category, ProductCategory as PC
        db.query(PC).filter(PC.product_id == product_id).delete()
        for cat_id in category_ids:
            cat = db.query(Category).filter(Category.id == cat_id).first()
            if cat:
                db.add(PC(product_id=product_id, category_id=cat_id))
    if collection_ids is not None:
        from app.models import Collection, ProductCollection as PColl
        db.query(PColl).filter(PColl.product_id == product_id).delete()
        for coll_id in collection_ids:
            coll = db.query(Collection).filter(Collection.id == coll_id).first()
            if coll:
                db.add(PColl(product_id=product_id, collection_id=coll_id))
    if category_ids is not None or collection_ids is not None:
        db.commit()
        db.refresh(updated)
    invalidate_stats()
    log_user(user, db, "update", "product", product_id, f"ویرایش محصول «{updated.name}»")
    return _enrich_product(updated, db)


class BulkProductAction(BaseModel):
    ids: list[int]
    set_active: bool | None = None
    set_collection_id: int | None = None
    clear_collections: bool = False
    set_tags: str | None = None
    set_notes: str | None = None
    add_category_ids: list[int] | None = None
    remove_category_ids: list[int] | None = None


@router.delete("/products/{product_id}")
def delete_product(product_id: int, user=Depends(require_any_role), db: Session = Depends(get_db)):
    repo = ProductRepository(db)
    if not repo.delete(product_id):
        raise HTTPException(status_code=404, detail="Product not found")
    invalidate_stats()
    return {"message": "Product deactivated", "id": product_id}


@router.post("/products/bulk")
def bulk_product_action(body: BulkProductAction, user=Depends(require_any_role), db: Session = Depends(get_db)):
    """Apply a single operation across many products at once."""
    if not body.ids:
        raise HTTPException(status_code=400, detail="هیچ محصولی انتخاب نشده است")

    products = db.query(Product).filter(Product.id.in_(body.ids)).all()
    changed = 0

    for p in products:
        if body.set_active is not None and p.is_active != body.set_active:
            p.is_active = body.set_active
            changed += 1
        if body.set_tags is not None:
            p.tags = body.set_tags
            changed += 1
        if body.set_notes is not None:
            p.notes = body.set_notes
            changed += 1

    # Collection assignment — replace all collection links for selected products
    if body.set_collection_id is not None or body.clear_collections:
        for p in products:
            db.query(ProductCollection).filter(ProductCollection.product_id == p.id).delete()
            if body.set_collection_id:
                coll = db.query(Collection).filter(Collection.id == body.set_collection_id).first()
                if coll:
                    db.add(ProductCollection(product_id=p.id, collection_id=coll.id))
        changed += len(products)

    # Category assignment — add/remove categories from selected products
    if body.add_category_ids or body.remove_category_ids:
        from app.models import Category, ProductCategory as PC
        add_ids = set(body.add_category_ids or [])
        remove_ids = set(body.remove_category_ids or [])
        for p in products:
            if add_ids:
                existing_ids = {pc.category_id for pc in db.query(PC).filter(PC.product_id == p.id).all()}
                for cat_id in add_ids - existing_ids:
                    cat = db.query(Category).filter(Category.id == cat_id, Category.is_active == True).first()
                    if cat:
                        db.add(PC(product_id=p.id, category_id=cat_id))
            if remove_ids:
                db.query(PC).filter(PC.product_id == p.id, PC.category_id.in_(remove_ids)).delete()
        changed += len(products)

    db.commit()
    invalidate_stats()
    log_user(user, db, "bulk", "product", summary=f"اعمال عملیات گروهی روی {len(products)} محصول")
    return {"updated": len(products), "changed": changed}


@router.delete("/products/{product_id}/permanent")
def permanent_delete_product(product_id: int, user=Depends(require_admin), db: Session = Depends(get_db)):
    from app.models import Product
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    db.delete(product)
    db.commit()
    invalidate_stats()
    return {"message": "Product permanently deleted", "id": product_id}


MAX_IMAGES = 6


@router.post("/products/{product_id}/images")
async def upload_product_images(
    product_id: int,
    files: list[UploadFile] = File(...),
    user=Depends(require_any_role),
    db: Session = Depends(get_db),
):
    """Upload one or more images for a product (max 5 total)."""
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Check current count
    current_count = db.query(ProductImage).filter(ProductImage.product_id == product_id).count()
    remaining = MAX_IMAGES - current_count
    if remaining <= 0:
        raise HTTPException(status_code=400, detail=f"حداکثر {MAX_IMAGES} تصویر مجاز است")
    if len(files) > remaining:
        raise HTTPException(status_code=400, detail=f"فقط {remaining} تصویر دیگر مجاز است")

    uploaded = []

    for file in files[:remaining]:
        content = await file.read()
        if len(content) > MAX_UPLOAD_SIZE:
            raise HTTPException(status_code=400, detail="حجم فایل نباید بیشتر از ۱۰ مگابایت باشد")

        ext = _validate_image_bytes(content, file.content_type or "")
        if not ext:
            raise HTTPException(status_code=400, detail="فایل نامعتبر است. فقط JPEG, PNG, WebP, GIF مجاز است")

        filename = _process_and_save_image(content, ext, UPLOAD_DIR)

        max_order = db.query(ProductImage).filter(ProductImage.product_id == product_id).count()
        img = ProductImage(
            product_id=product_id,
            image_url=f"/uploads/{filename}",
            sort_order=max_order,
            is_primary=(current_count == 0 and len(uploaded) == 0),
        )
        db.add(img)
        uploaded.append(img)

    # Sync primary to product.image_url
    primary = db.query(ProductImage).filter(
        ProductImage.product_id == product_id, ProductImage.is_primary == True
    ).first()
    if primary:
        product.image_url = primary.image_url
    elif uploaded:
        product.image_url = uploaded[0].image_url

    try:
        db.commit()
    except Exception as err:
        logger.exception("Failed to commit product image upload: %s", err)
        db.rollback()
        raise HTTPException(status_code=500, detail="خطا در ذخیره تصاویر")

    for img in uploaded:
        db.refresh(img)

    return {
        "message": f"{len(uploaded)} تصویر آپلود شد",
        "images": [
            {"id": img.id, "image_url": img.image_url, "sort_order": img.sort_order, "is_primary": img.is_primary}
            for img in db.query(ProductImage).filter(ProductImage.product_id == product_id).order_by(ProductImage.sort_order).all()
        ],
    }


@router.delete("/products/{product_id}/images/{image_id}")
def delete_product_image_by_id(product_id: int, image_id: int, user=Depends(require_any_role), db: Session = Depends(get_db)):
    """Delete a specific image from a product."""
    img = db.query(ProductImage).filter(
        ProductImage.id == image_id, ProductImage.product_id == product_id
    ).first()
    if not img:
        raise HTTPException(status_code=404, detail="تصویر یافت نشد")

    # Delete file from disk
    filename = img.image_url.split("/")[-1]
    filepath = os.path.join(UPLOAD_DIR, filename)
    if os.path.exists(filepath):
        os.remove(filepath)

    was_primary = img.is_primary
    db.delete(img)
    db.flush()

    # If deleted was primary, assign new primary
    product = db.query(Product).filter(Product.id == product_id).first()
    if was_primary and product:
        new_primary = db.query(ProductImage).filter(
            ProductImage.product_id == product_id
        ).order_by(ProductImage.sort_order).first()
        if new_primary:
            new_primary.is_primary = True
            product.image_url = new_primary.image_url
        else:
            product.image_url = None

    try:
        db.commit()
    except Exception as err:
        logger.exception("Failed to delete product image: %s", err)
        db.rollback()
        raise HTTPException(status_code=500, detail="خطا در حذف تصویر")

    return {
        "message": "تصویر حذف شد",
        "images": [
            {"id": i.id, "image_url": i.image_url, "sort_order": i.sort_order, "is_primary": i.is_primary}
            for i in db.query(ProductImage).filter(ProductImage.product_id == product_id).order_by(ProductImage.sort_order).all()
        ],
    }


@router.put("/products/{product_id}/images/{image_id}/primary")
def set_primary_image(product_id: int, image_id: int, user=Depends(require_any_role), db: Session = Depends(get_db)):
    """Set an image as the primary (catalog display) image."""
    img = db.query(ProductImage).filter(
        ProductImage.id == image_id, ProductImage.product_id == product_id
    ).first()
    if not img:
        raise HTTPException(status_code=404, detail="تصویر یافت نشد")

    # Unset all primaries for this product
    db.query(ProductImage).filter(
        ProductImage.product_id == product_id
    ).update({"is_primary": False})

    img.is_primary = True
    product = db.query(Product).filter(Product.id == product_id).first()
    if product:
        product.image_url = img.image_url

    try:
        db.commit()
    except Exception as err:
        logger.exception("Failed to set primary image: %s", err)
        db.rollback()
        raise HTTPException(status_code=500, detail="خطا در تنظیم تصویر اصلی")

    return {
        "message": "تصویر اصلی تنظیم شد",
        "images": [
            {"id": i.id, "image_url": i.image_url, "sort_order": i.sort_order, "is_primary": i.is_primary}
            for i in db.query(ProductImage).filter(ProductImage.product_id == product_id).order_by(ProductImage.sort_order).all()
        ],
    }


@router.put("/products/{product_id}/images/reorder")
def reorder_images(product_id: int, body: ImageReorderRequest, user=Depends(require_any_role), db: Session = Depends(get_db)):
    """Reorder images. Body: { "order": [image_id, image_id, ...] }"""
    order = body.order
    if not order:
        raise HTTPException(status_code=400, detail="لیست ترتیب الزامی است")

    images = db.query(ProductImage).filter(
        ProductImage.product_id == product_id,
        ProductImage.id.in_(order),
    ).all()
    img_map = {img.id: img for img in images}

    for idx, img_id in enumerate(order):
        if img_id in img_map:
            img_map[img_id].sort_order = idx

    try:
        db.commit()
    except Exception as err:
        logger.exception("Failed to reorder images: %s", err)
        db.rollback()
        raise HTTPException(status_code=500, detail="خطا در مرتب‌سازی تصاویر")

    return {
        "message": "ترتیب به‌روز شد",
        "images": [
            {"id": i.id, "image_url": i.image_url, "sort_order": i.sort_order, "is_primary": i.is_primary}
            for i in db.query(ProductImage).filter(ProductImage.product_id == product_id).order_by(ProductImage.sort_order).all()
        ],
    }


# ── Dimension extraction from 3MF/STL ────────────────────────────────────

def _extract_dimensions_from_3mf(file_path: str) -> dict:
    """Extract bounding box dimensions from a 3MF file."""
    import zipfile
    from xml.etree.ElementTree import fromstring
    
    try:
        with zipfile.ZipFile(file_path, 'r') as zf:
            for name in zf.namelist():
                if name.endswith('.model'):
                    data = zf.read(name)
                    root = fromstring(data)
                    # Find all vertices
                    ns = {'': 'http://schemas.microsoft.com/3dmanufacturing/core/2015/02'}
                    vertices = root.findall('.//{http://schemas.microsoft.com/3dmanufacturing/core/2015/02}vertex')
                    if not vertices:
                        vertices = root.findall('.//vertex')
                    
                    min_x = min_y = min_z = float('inf')
                    max_x = max_y = max_z = float('-inf')
                    
                    for v in vertices:
                        x = float(v.get('x', 0))
                        y = float(v.get('y', 0))
                        z = float(v.get('z', 0))
                        min_x, max_x = min(min_x, x), max(max_x, x)
                        min_y, max_y = min(min_y, y), max(max_y, y)
                        min_z, max_z = min(min_z, z), max(max_z, z)
                    
                    if min_x != float('inf'):
                        return {
                            "dimension_x": round(max_x - min_x, 2),
                            "dimension_y": round(max_y - min_y, 2),
                            "dimension_z": round(max_z - min_z, 2),
                        }
    except Exception as err:
        logger.warning("Could not extract dimensions from 3MF %s: %s", file_path, err)
    return {}


def _extract_dimensions_from_stl(file_path: str) -> dict:
    """Extract bounding box dimensions from an STL file (binary or ASCII)."""
    import struct
    
    try:
        with open(file_path, 'rb') as f:
            header = f.read(80)
            # Binary STL
            if header[:5] == b'solid' and b'facet' not in header[:100]:
                num_triangles = struct.unpack('<I', f.read(4))[0]
                min_x = min_y = min_z = float('inf')
                max_x = max_y = max_z = float('-inf')
                
                for _ in range(num_triangles):
                    f.read(12)  # normal
                    for _ in range(3):
                        x, y, z = struct.unpack('<fff', f.read(12))
                        min_x, max_x = min(min_x, x), max(max_x, x)
                        min_y, max_y = min(min_y, y), max(max_y, y)
                        min_z, max_z = min(min_z, z), max(max_z, z)
                    f.read(2)  # attribute
                
                if min_x != float('inf'):
                    return {
                        "dimension_x": round(max_x - min_x, 2),
                        "dimension_y": round(max_y - min_y, 2),
                        "dimension_z": round(max_z - min_z, 2),
                    }
    except Exception as e:
        logging.getLogger(__name__).warning("STL dimension extraction failed: %s", e)
    return {}


@router.post("/products/{product_id}/dimensions")
def extract_dimensions(product_id: int, db: Session = Depends(get_db)):
    """Extract dimensions from product's 3MF/STL model file."""
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    if not product.model_file:
        raise HTTPException(status_code=400, detail="فایل مدلی برای این محصول وجود ندارد")
    
    # Resolve file path
    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "uploads")
    model_path = os.path.join(uploads_dir, product.model_file)
    
    if not os.path.isfile(model_path):
        raise HTTPException(status_code=404, detail="فایل مدل یافت نشد")
    
    # Extract dimensions based on file type
    ext = os.path.splitext(model_path)[1].lower()
    dims = {}
    if ext == '.3mf':
        dims = _extract_dimensions_from_3mf(model_path)
    elif ext == '.stl':
        dims = _extract_dimensions_from_stl(model_path)
    
    if not dims:
        raise HTTPException(status_code=400, detail="خطا در استخراج ابعاد از فایل مدل")
    
    # Update product with dimensions
    product.dimension_x = dims.get('dimension_x')
    product.dimension_y = dims.get('dimension_y')
    product.dimension_z = dims.get('dimension_z')
    db.commit()
    db.refresh(product)
    
    return {
        "message": "ابعاد با موفقیت استخراج شد",
        "dimensions": dims,
    }


from app.routers.auth import limiter, require_admin, Request


# ── Stateless calculator ──────────────────────────────────────────────

@router.post("/calculate", response_model=CalculateResponse)
@limiter.limit("30/minute")
def stateless_calculator(request: Request, req: CalculateRequest, db: Session = Depends(get_db)):
    """Calculate costs without creating a product."""
    costs = calculate_product_costs(
        db,
        weight_g=req.weight_g,
        support_g=req.support_g,
        flushed_g=req.flushed_g,
        print_time_hours=req.print_time_hours,
        post_pro_hours=req.post_pro_hours,
        extras_cost=req.extras_cost,
        machine_id=req.machine_id,
        material_id=req.material_id,
    )
    return costs
