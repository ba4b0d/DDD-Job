#!/usr/bin/env python3
"""
Import 3DJAT product data from an Excel (.xlsx) file.

Usage:
    python -m app.import_excel path/to/file.xlsx

This script reads the Excel spreadsheet and upserts data into the SQLite DB.
Run manually when you have updated spreadsheet data.
"""
import sys
import os

# Add parent dir to path so we can import app.*
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip install openpyxl")
    sys.exit(1)

from app.database import SessionLocal, engine, Base
from app.models import Settings, Machine, Material, Product


def parse_time(time_str: str) -> float:
    """Convert H:MM:SS or MM:SS string to decimal hours."""
    if not time_str or not isinstance(time_str, str):
        return 0.0
    parts = time_str.strip().split(":")
    try:
        if len(parts) == 3:
            h, m, s = parts
            return int(h) + int(m) / 60 + int(s) / 3600
        elif len(parts) == 2:
            h, m = parts
            return int(h) + int(m) / 60
        else:
            return float(time_str)
    except (ValueError, TypeError):
        return 0.0


def safe_float(val, default=0.0) -> float:
    """Safely convert value to float."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_str(val, default="") -> str:
    """Safely convert value to string."""
    if val is None:
        return default
    return str(val).strip()


def import_excel(filepath: str):
    """Import data from Excel file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    db = SessionLocal()

    # Build lookups
    machine_map = {m.name: m.id for m in db.query(Machine).all()}
    mat_lookup = {(m.name, m.color): m.id for m in db.query(Material).all()}

    imported = 0
    skipped = 0

    # Try to find the products sheet (first sheet or named)
    sheet = wb.active

    # Assume row 1 is headers, data starts at row 2
    headers = {}
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val:
            headers[str(val).strip().lower()] = col

    print(f"Found headers: {list(headers.keys())}")

    for row in range(2, sheet.max_row + 1):
        # Read cells by header name
        def cell(name):
            col = headers.get(name.lower())
            if col:
                return sheet.cell(row=row, column=col).value
            return None

        name = safe_str(cell("name"))
        if not name:
            skipped += 1
            continue

        product_id = safe_str(cell("product_id") or cell("id") or "")
        qty = int(safe_float(cell("qty"), 1))
        machine_name = safe_str(cell("machine") or cell("machine_name") or "")
        material_name = safe_str(cell("material") or cell("material_name") or "")
        color = safe_str(cell("color") or "")
        weight_g = safe_float(cell("weight_g") or cell("weight") or 0)
        support_g = safe_float(cell("support_g") or cell("support") or 0)
        flushed_g = safe_float(cell("flushed_g") or cell("flushed") or 0)
        print_time_hours = safe_float(cell("print_time_hours") or 0)
        post_pro_hours = safe_float(cell("post_pro_hours") or 0)
        extras_cost = safe_float(cell("extras_cost") or 0)
        final_price = cell("final_price")
        category = safe_str(cell("category") or "")
        notes = safe_str(cell("notes") or "")

        # Look up machine and material IDs
        machine_id = machine_map.get(machine_name)
        material_id = mat_lookup.get((material_name, color))

        # Skip if required references missing
        if not machine_id:
            print(f"  Row {row}: Unknown machine '{machine_name}', skipping")
            skipped += 1
            continue
        if not material_id:
            print(f"  Row {row}: Unknown material '{material_name}'/'{color}', skipping")
            skipped += 1
            continue

        existing = db.query(Product).filter(
            Product.name == name,
            Product.product_id == product_id,
        ).first()

        if existing:
            existing.qty = qty
            existing.machine_id = machine_id
            existing.material_id = material_id
            existing.weight_g = weight_g
            existing.support_g = support_g
            existing.flushed_g = flushed_g
            existing.print_time_hours = print_time_hours
            existing.post_pro_hours = post_pro_hours
            existing.extras_cost = extras_cost
            existing.final_price = safe_float(final_price) if final_price else None
            existing.category = category
            existing.notes = notes
        else:
            db.add(Product(
                product_id=product_id,
                name=name,
                qty=qty,
                machine_id=machine_id,
                material_id=material_id,
                weight_g=weight_g,
                support_g=support_g,
                flushed_g=flushed_g,
                print_time_hours=print_time_hours,
                post_pro_hours=post_pro_hours,
                extras_cost=extras_cost,
                final_price=safe_float(final_price) if final_price else None,
                category=category,
                notes=notes,
            ))

        imported += 1

    db.commit()
    db.close()
    print(f"Import complete: {imported} products imported, {skipped} skipped")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python -m app.import_excel <path_to_excel_file.xlsx>")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        sys.exit(1)

    # Ensure tables exist
    Base.metadata.create_all(bind=engine)
    import_excel(filepath)
